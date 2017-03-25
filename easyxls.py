import sys, getopt, os
import openpyxl

######################################################################################################################

def get_column_based_header_fields(max_column, header_row_start, header_col_start, ws):

	header_array = {}

	col_id = header_col_start
	
	while ord(col_id) <= ord(max_column):
		try:                    
			column_string = col_id + str(header_row_start)
			val = str(ws[column_string].value)
			header_array[col_id] = val
			col_id = chr(ord(col_id) + 1)
		except:
			break
	# End While

	return header_array
	
# End Function
######################################################################################################################

def get_row_based_header_fields(max_row, header_row_start, header_col_start, ws):

	header_array = {}

	row_num = int(header_row_start)
	
	while row_num <= int(max_row):
		try:
			column_string = header_col_start + str(row_num)
			val = str(ws[column_string].value)
			header_array[row_num] = val
			row_num = row_num + 1
		except:
			break
	# End While

	return header_array
	
# End Function
######################################################################################################################

def get_with_column_headings(spreadsheet_array, max_row, max_column, header_row_start, header_col_start, ws):

	anonymous_header_count = 0

	num_rows = ws.max_row
	row_num = int(header_row_start) + 1

	header_array = get_column_based_header_fields(max_column, header_row_start, header_col_start, ws)
	finished = 0

	while finished != 1:
		col_id = header_col_start
		new_dict = {}
		
		while ord(col_id)  <= ord(max_column):
			try:                    
				column_string = col_id + str(row_num)
				val = str(ws[column_string].value)

				if header_array[col_id] != 'None':
					new_dict[header_array[col_id]] = val
				else:
					new_dict[anonymous_header_count] = val
					anonymous_header_count = anonymous_header_count + 1
				# End If

				col_id = chr(ord(col_id) + 1)
			except:
				break
		# End While
	
		spreadsheet_array.append(new_dict)

		if row_num >= num_rows or row_num > max_row:
			finished = 1
		else:
			row_num = row_num + 1
		# End If

	# End While
	
	return spreadsheet_array
# End Function
######################################################################################################################

def get_with_row_headings(spreadsheet_array, max_row, max_column, header_row_start, header_col_start, ws):

	anonymous_header_count = 0

	num_rows = ws.max_row
	col_id = chr(ord(header_col_start) + 1)

	header_array = get_row_based_header_fields(max_row, header_row_start, header_col_start, ws)
	finished = 0

	while finished != 1:
		new_dict = {}
		row_num = int(header_row_start)

		while row_num <= max_row:
			try:                    
				column_string = col_id + str(row_num)
				val = str(ws[column_string].value)
				
				if header_array[row_num ] != 'None':
					new_dict[header_array[row_num]] = val
				else:
					new_dict[anonymous_header_count] = val
					anonymous_header_count = anonymous_header_count + 1
				# End If

				row_num = row_num + 1
			except:
				break
		# End While

		spreadsheet_array.append(new_dict)
		
		if ord(col_id) >= ord(max_column):
			finished = 1
		else:
			col_id = chr(ord(col_id) + 1)
		# End If

	# End While
	return spreadsheet_array
# End Function
######################################################################################################################

def get_with_pivot_table(spreadsheet_dict, max_row, max_column, header_row_start, header_col_start, ws):
	
	anonymous_header_count = 0
	
	col_header_array = get_column_based_header_fields(max_column, header_row_start, header_col_start, ws)
	row_header_array = get_row_based_header_fields(max_row, header_row_start, header_col_start, ws)
	
	row_num = int(header_row_start) + 1
	
	while row_num <= int(max_row):
		
		current_top_level_key = ""
		
		if row_header_array[row_num] != 'None':
			current_top_level_key = row_header_array[row_num]
			spreadsheet_dict[current_top_level_key] = {}
		else:
			current_top_level_key = str(anonymous_header_count)
			spreadsheet_dict[current_top_level_key] = {}
			anonymous_header_count = anonymous_header_count + 1
		# End If
		
		col_id = chr(ord(header_col_start) + 1)
		
		while ord(col_id) <= ord(max_column):
			column_string = col_id + str(row_num)
			val = str(ws[column_string].value)
			
			if col_header_array[col_id] != 'None':
				spreadsheet_dict[current_top_level_key] [col_header_array[col_id]] = val
			else:
				spreadsheet_dict[current_top_level_key] [anonymous_header_count] = val
				anonymous_header_count = anonymous_header_count + 1
			# End If
		
			col_id = chr(ord(col_id) + 1)
		# End While
		
		row_num = row_num + 1
	# End While
	
	return spreadsheet_dict
# End Function
######################################################################################################################
######################################################################################################################

def get_spreadsheet(spreadsheet_path, max_row, max_column, header_row_start, header_col_start, format):

	global spreadsheet_array
	global spreadsheet_dict

	spreadsheet_array = []
	spreadsheet_dict = {}
	
	wb = openpyxl.load_workbook(filename =spreadsheet_path, data_only = True)
	ws = wb.get_sheet_by_name(name = 'Sheet1')

	print "Spreadsheet has been read into memory...\n"

	if format == "row":
		get_with_row_headings(spreadsheet_array, max_row, max_column, header_row_start, header_col_start, ws)
		return spreadsheet_array
		
	elif format == "column":
		get_with_column_headings(spreadsheet_array, max_row, max_column, header_row_start, header_col_start, ws)
		return spreadsheet_array

	elif format == "pivot":
		get_with_pivot_table(spreadsheet_dict, max_row, max_column, header_row_start, header_col_start, ws)
		return spreadsheet_dict
		
	# End If
	
# End Function
######################################################################################################################
